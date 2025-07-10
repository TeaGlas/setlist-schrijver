from fastapi import FastAPI, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
import random
from openpyxl import load_workbook
from io import BytesIO

app = FastAPI()

@app.get("/")
def read_root():
    return {"message": "API is running!"}

# Allow frontend (e.g., GitHub Pages) to access
app.add_middleware(
    CORSMiddleware,
    # Set specific domain in production!
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/generate")
async def generate_songs(
        file: UploadFile,
        total_songs: int = Form(...),
        min_dutch: int = Form(0),
        min_german: int = Form(0),
        min_english: int = Form(0),
        include_christmas: bool = Form(True),
        christmas_count: int = Form(0),
        booklet_option: str = Form("any"),
        exclude_song_numbers: Optional[str] = Form(""),
):
    # Load workbook
    contents = await file.read()
    wb = load_workbook(filename=BytesIO(contents), data_only=True)
    sheetnames = wb.sheetnames
    if not sheetnames:
        return {"error": "No sheets found in uploaded Excel file."}
    sheet = wb[sheetnames[0]]

    # Assume headers: Number, Title, Language, Booklet, Christmas
    songs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        songs.append({
            "number": row[0],
            "title": row[1],
            "artist": row[2],
            "language": row[3],
            "christmas": str(row[4]).strip().lower() in ["true", "1", "yes"],
            "booklet": row[5],
        })

    # Exclude songs
    excluded_numbers = set()
    if exclude_song_numbers:
        excluded_numbers = set(map(int, exclude_song_numbers.split(",")))
        songs = [s for s in songs if s["number"] not in excluded_numbers]

    # Christmas filter
    if not include_christmas:
        songs = [s for s in songs if not s["christmas"]]

    def pop_n_random(language: str, n: int):
        selected = [
            s for s in songs if s["language"].lower() == language.lower()
        ]
        picked = random.sample(selected, min(n, len(selected)))
        for song in picked:
            songs.remove(song)
        return picked

    result = []
    result += pop_n_random("Dutch", min_dutch)
    result += pop_n_random("English", min_english)
    result += pop_n_random("German", min_german)

    if include_christmas and christmas_count:
        christmas_songs = [s for s in songs if s["christmas"]]
        picked = random.sample(christmas_songs,
                               min(christmas_count, len(christmas_songs)))
        for song in picked:
            songs.remove(song)
        result += picked

    # Booklet filtering
    if booklet_option.startswith("only:"):
        chosen_booklet = booklet_option.split(":")[1]
        songs = [s for s in songs if str(s["booklet"]) == chosen_booklet]
    elif booklet_option == "one-from-each":
        all_booklets = set(s["booklet"] for s in songs)
        for b in all_booklets:
            from_booklet = [s for s in songs if s["booklet"] == b]
            if from_booklet:
                chosen = random.choice(from_booklet)
                result.append(chosen)
                songs.remove(chosen)

    # Fill up remaining songs
    remaining = total_songs - len(result)
    if remaining > 0:
        result += random.sample(songs, min(remaining, len(songs)))

    return {"songs": result}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000)
